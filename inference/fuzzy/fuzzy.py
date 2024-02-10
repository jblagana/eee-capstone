import matplotlib.pyplot as plt
from fuzzy_expert.variable import FuzzyVariable
from fuzzy_expert.rule import FuzzyRule
from fuzzy_expert.inference import DecompositionalInference
from itertools import product
from ipywidgets import interact, widgets
import openpyxl
import time

def generate_variables():
    variables = {
        "cover": FuzzyVariable(universe_range=(0, 100), terms={
            "Low(rarely hidden faces)": [(20, 1), (40, 0)],
            "Medium(occasional hidden faces)": [(20, 0), (50, 1), (80, 0)],
            "High(prolonged hidden faces)": [(60, 0), (80, 1)],
        }),
        "crowd": FuzzyVariable(universe_range=(0, 100), terms={
            "Low(highly crowded)": [(20, 1), (40, 0)],
            "Medium(moderately crowded)": [(20, 0), (50, 1), (80, 0)],
            "High(not crowded)": [(60, 0), (80, 1)],
        }),
        "loiter": FuzzyVariable(universe_range=(0, 100), terms={
            "Low(minimal loitering)": [(20, 1), (40, 0)],
            "Medium(moderate loitering)": [(20, 0), (50, 1), (80, 0)],
            "High(heavy loitering)": [(60, 0), (80, 1)],
        }),
        "decision": FuzzyVariable(universe_range=(0, 100), terms={
            "Very-Low(very low robbery risk)": [(10, 1), (30, 0)],
            "Low(low robbery risk)": [(20, 0), (35, 1), (50, 0)],
            "Medium(moderate robbery risk)": [(35, 0), (50, 1), (65, 0)],
            "High(high robbery risk)": [(50, 0), (65, 1), (80, 0)],
            "Very-High(very high robbery risk)": [(70, 0), (90, 1)],
        }),
    }
    return variables

def generate_rules(variables):
    linguistic_combo = product(*[var[1].terms.keys() for var in variables.items() if var[0] != "decision"])
    linguistic_combo = list(linguistic_combo)
    
    decision_terms = list([var[1].terms.keys() for var in variables.items() if var[0] == "decision"][0])
    decision_list = [x for x in decision_terms for _ in range(len(linguistic_combo)//len(decision_terms))]
    
    while len(decision_list) < len(linguistic_combo):
        decision_list.append(decision_terms[-1])

    temp = [(x[0], x[1], x[2], y) for x, y in zip(linguistic_combo, decision_list)]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = list(variables.keys())
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    for row_num, combo in enumerate(temp, start=2):
        for col_num, value in enumerate(combo + ("",), start=1):
            sheet.cell(row=row_num, column=col_num, value=value)

    workbook.save("combo.xlsx")

    rules = []
    for i, combo in enumerate(linguistic_combo):
        premise = [(variable, term) if j == 0 else ("AND", variable, term) for j, (variable, term) in enumerate(zip(variables.keys(), combo))]
        consequence = [("decision", decision_list[i])]
        rules.append(FuzzyRule(premise=premise, consequence=consequence))

    return rules

def demo(cover, crowd, loiter):
    model = DecompositionalInference(
        and_operator="min",
        or_operator="max",
        implication_operator="Rc",
        composition_operator="max-min",
        production_link="max",
        defuzzification_operator="cog",
    )

    plt.figure(figsize=(10, 6))
    model.plot(
        variables=variables,
        rules=rules,
        cover=cover,
        crowd=crowd,
        loiter=loiter,
    )

# Generate variables and rules
variables = generate_variables()
rules = generate_rules(variables)

# interact(
#     demo,
#     cover=widgets.FloatSlider(min=0, max=100),
#     crowd=widgets.FloatSlider(min=0, max=100),
#     loiter=widgets.FloatSlider(min=0, max=100),
# )


# Benchmarking function
def benchmark_inference_speed(num_samples=100):
    cover_values = [i for i in range(0, 101, 100 // num_samples)]
    crowd_values = [i for i in range(0, 101, 100 // num_samples)]
    loiter_values = [i for i in range(0, 101, 100 // num_samples)]

    inference_times = []

    for cover, crowd, loiter in product(cover_values, crowd_values, loiter_values):
        start_time = time.time()
        model = DecompositionalInference(
            and_operator="min",
            or_operator="max",
            implication_operator="Rc",
            composition_operator="max-min",
            production_link="max",
            defuzzification_operator="cog",
        )
        model(variables=variables, rules=rules, cover=cover, crowd=crowd, loiter=loiter)
        end_time = time.time()
        inference_times.append(end_time - start_time)
        print(cover, crowd, loiter)

    plt.figure(figsize=(10, 6))
    plt.plot(inference_times, label="Inference Time")
    plt.xlabel("Sample Index")
    plt.ylabel("Time (seconds)")
    plt.title("Inference Speed Benchmark")
    plt.legend()
    plt.show()

# Uncomment the line below to run the benchmark
benchmark_inference_speed(2)